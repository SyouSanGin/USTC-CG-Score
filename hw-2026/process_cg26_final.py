# -*- coding: utf-8 -*-
"""
CG26 Final Team Formation Statistics Generator
==============================================
Inputs:
  - CG26 Final 组队统计.xlsx : 38 teams with captain, members, topic
  - StudentNames-new.xlsx   : ~80 valid students with IDs, grades, departments

Outputs:
  - CG26FinalProject组队统计（最终版）.xlsx : 4 sheets with full analysis
  - team_summary_cg26.json              : structured summary
"""
import pandas as pd
import json
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG
# ============================================================
ROSTER_PATH = r'D:\Projects\USTC-CG-Score\hw-2026\StudentNames-new.xlsx'
TEAMS_PATH = r'D:\Projects\USTC-CG-Score\hw-2026\CG26 Final 组队统计.xlsx'
OUTPUT_XLSX = r'D:\Projects\USTC-CG-Score\hw-2026\CG26FinalProject组队统计（最终版）.xlsx'
OUTPUT_JSON = r'D:\Projects\USTC-CG-Score\hw-2026\team_summary_cg26.json'

# Number of topic categories for analysis (explicitly counted, not estimated)
# These will be computed from actual data

# ============================================================
# STEP 1: Load Roster
# ============================================================
def load_roster(path):
    """Load student roster. Returns name_to_info dict and set of valid names."""
    df = pd.read_excel(path)
    df = df[df['姓名'].notna()].copy()
    df = df.dropna(subset=['学号'])

    name_to_info = {}
    for _, row in df.iterrows():
        roster_name = str(row['姓名']).strip()
        clean_name = roster_name.rstrip('*')
        name_to_info[clean_name] = {
            'roster_id': int(row['ID号']),
            'student_id': str(row['学号']).strip(),
            'roster_name': roster_name,  # original name with possible *
            'grade': int(row['当前年级']) if pd.notna(row['当前年级']) else None,
            'department': str(row['当前院系']).strip() if pd.notna(row['当前院系']) else '',
            'class_name': str(row['行政班']).strip() if pd.notna(row['行政班']) else '',
            'has_asterisk': roster_name.endswith('*'),
        }

    valid_names = set(name_to_info.keys())
    print(f"[ROSTER] Total valid students: {len(valid_names)}")
    asterisk_count = sum(1 for v in name_to_info.values() if v['has_asterisk'])
    print(f"[ROSTER] Names with trailing '*': {asterisk_count}")

    # Check for collisions after stripping *
    stripped_names = defaultdict(list)
    for name in valid_names:
        stripped_names[name].append(name)
    collisions = {k: v for k, v in stripped_names.items() if len(v) > 1}
    if collisions:
        print(f"[ROSTER] WARNING: Duplicate names after stripping '*': {collisions}")

    return name_to_info, valid_names


def load_teams(path):
    """Load team formation stats. Returns list of raw team dicts."""
    df = pd.read_excel(path)
    print(f"[TEAMS] Loaded {len(df)} rows, columns: {list(df.columns)}")

    raw_teams = []
    for idx, row in df.iterrows():
        captain = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
        m2 = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
        m3 = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
        recorded_size = int(row.iloc[4]) if pd.notna(row.iloc[4]) else 0
        topic = str(row.iloc[5]).strip() if pd.notna(row.iloc[5]) else ''
        team_num = int(row.iloc[0]) if pd.notna(row.iloc[0]) else 0

        # Build members list (filter NaN placeholders)
        members_raw = [captain]
        if m2 and m2.lower() not in ('nan', 'none', ''):
            members_raw.append(m2)
        if m3 and m3.lower() not in ('nan', 'none', ''):
            members_raw.append(m3)

        raw_teams.append({
            'row_idx': idx,
            'team_num': team_num,  # original team number from the file
            'captain': captain,
            'members_raw': members_raw,
            'recorded_size': recorded_size,
            'topic': topic if topic and topic.lower() != 'nan' else '',
        })

    # Summary
    sizes = Counter(len(t['members_raw']) for t in raw_teams)
    print(f"[TEAMS] Parsed {len(raw_teams)} teams: 3p={sizes.get(3,0)}, 2p={sizes.get(2,0)}, 1p={sizes.get(1,0)}")

    # Detect size discrepancies
    for t in raw_teams:
        actual = len(t['members_raw'])
        if actual != t['recorded_size']:
            print(f"[TEAMS] DISCREPANCY: Team {t['team_num']} (row {t['row_idx']}) "
                  f"recorded_size={t['recorded_size']}, actual members={actual}")

    return raw_teams


# ============================================================
# STEP 2: Validate & Cross-Reference
# ============================================================
def validate_teams(raw_teams, valid_names):
    """Validate team members against roster. Handle duplicates, invalid names."""
    all_submitted = set()
    for t in raw_teams:
        all_submitted.update(t['members_raw'])

    invalid_names = all_submitted - valid_names
    valid_submitted = all_submitted & valid_names

    print(f"\n[VALIDATE] Names in teams: {len(all_submitted)}")
    print(f"[VALIDATE] Valid matches: {len(valid_submitted)}")
    print(f"[VALIDATE] Invalid (not in roster): {len(invalid_names)}")
    if invalid_names:
        print(f"[VALIDATE] Invalid names: {sorted(invalid_names)}")

    # Build validated teams (deduplicate students across teams, first-wins)
    student_assigned = set()
    validated_teams = []
    duplicates_found = []

    for t in raw_teams:
        valid_members = []
        invalid_members = []
        for m in t['members_raw']:
            if m in valid_names:
                valid_members.append(m)
            else:
                invalid_members.append(m)

        # Deduplicate within team
        seen = set()
        deduped = []
        for m in valid_members:
            if m not in seen:
                seen.add(m)
                deduped.append(m)
        valid_members = deduped

        if len(valid_members) == 0:
            print(f"[VALIDATE] Team {t['team_num']}: ALL members invalid! Raw: {t['members_raw']}")
            continue

        # Check cross-team duplicates
        cleaned = []
        for m in valid_members:
            if m in student_assigned:
                duplicates_found.append(f"  {m}: already in another team (appears in team {t['team_num']})")
            else:
                cleaned.append(m)
                student_assigned.add(m)

        if len(cleaned) == 0:
            print(f"[VALIDATE] Team {t['team_num']}: All members already assigned to other teams")
            continue

        # Detect size discrepancy
        notes = []
        if len(t['members_raw']) != t['recorded_size']:
            notes.append(f"人数列记载{t['recorded_size']}人,实际{len(t['members_raw'])}人")
        if invalid_members:
            notes.append(f"无效成员: {', '.join(invalid_members)}")
        if len(valid_members) != len(t['members_raw']):
            notes.append(f"原始{t['members_raw']} -> 有效{cleaned}")

        validated_teams.append({
            'team_num': t['team_num'],
            'members': cleaned,
            'size': len(cleaned),
            'topic': t['topic'] if t['topic'] else '未确定',
            'captain': t['captain'],
            'notes': '; '.join(notes),
            'invalid_members': invalid_members,
            'has_invalid': len(invalid_members) > 0,
        })

    if duplicates_found:
        print(f"[VALIDATE] Cross-team duplicates detected:")
        for d in duplicates_found:
            print(d)

    return validated_teams, invalid_names, student_assigned


def compute_unpaired(valid_names, student_assigned):
    """Find students in roster but not in any team."""
    unpaired = valid_names - student_assigned
    print(f"\n[UNPAIRED] Students not in any team: {len(unpaired)}")
    return unpaired


# ============================================================
# STEP 3: Topic Categorization
# ============================================================
TOPIC_CAT_UNKNOWN = '未确定'
TOPIC_CAT_GAME = '游戏开发'
TOPIC_CAT_RENDER = '渲染与模拟'
TOPIC_CAT_AI = 'AI/LLM相关'
TOPIC_CAT_MOD = '游戏MOD'
TOPIC_CAT_OTHER = '其他方向'

def categorize_topic(topic):
    """Categorize a topic string into a high-level category."""
    if not topic or topic in ('未确定', '未定', '未知', '还没确定', '还没想好', '没想好', '还没想呢'):
        return TOPIC_CAT_UNKNOWN
    t = topic.lower()
    # AI/LLM first so LLM-based modeling doesn't get caught by "建模"
    if any(kw in t for kw in ('llm', 'ai', 'aigc', 'gpt', '大模型')):
        return TOPIC_CAT_AI
    if any(kw in t for kw in ('游戏', 'unity', 'fps', '跑酷', 'boxel', 'game')):
        return TOPIC_CAT_GAME
    if any(kw in t for kw in ('mod', 'outer wilds')):
        return TOPIC_CAT_MOD
    if any(kw in t for kw in ('渲染', '模拟', '仿真', 'shader', '光影', '光追', 'blender', 'opengl', 'mesh', 'reduction', '建模', '3d')):
        return TOPIC_CAT_RENDER
    if any(kw in t for kw in ('cad', )):
        return TOPIC_CAT_RENDER
    return TOPIC_CAT_OTHER


# ============================================================
# Excel Styles
# ============================================================
FONT = Font(name='微软雅黑', size=11)
FONT_SMALL = Font(name='微软雅黑', size=10)
FONT_BOLD = Font(name='微软雅黑', bold=True, size=11)
FONT_BOLD_LARGE = Font(name='微软雅黑', bold=True, size=13)
FONT_HEADER = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
FONT_GREEN = Font(name='微软雅黑', size=11, color='006100')
FONT_RED = Font(name='微软雅黑', size=11, color='9C0006')
FONT_ORANGE = Font(name='微软雅黑', size=11, color='BF8F00')
FONT_BLUE = Font(name='微软雅黑', size=11, color='305496')

FILL_HEADER = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
FILL_3 = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Green
FILL_2 = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')  # Blue
FILL_1 = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')  # Orange
FILL_SUMMARY = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
FILL_RED_LIGHT = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
FILL_SECTION = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
FILL_WHITE = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

FILL_BY_SIZE = {3: FILL_3, 2: FILL_2, 1: FILL_1}
LABEL_BY_SIZE = {3: '3人组', 2: '2人组', 1: '1人组'}


def style_header_row(ws, row, cols, fill=None):
    """Apply header styling to a row."""
    if fill is None:
        fill = FILL_HEADER
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_HEADER
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_cell(ws, row, col, value, font=None, fill=None, align=None, number_format=None):
    """Write and style a single cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or FONT
    cell.alignment = align or CENTER
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format


def write_section_header(ws, row, cols, text, fill=None):
    """Write a section header spanning multiple columns."""
    fill = fill or FILL_SECTION
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = FONT_BOLD
    cell.fill = fill
    cell.alignment = LEFT
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).border = THIN_BORDER
        ws.cell(row=row, column=c).fill = fill


# ============================================================
# Sheet 1: 最终组队统计
# ============================================================
def generate_sheet1(wb, all_sorted_teams, name_to_info):
    """Sheet 1: Final team stats with member details and color coding."""
    ws = wb.create_sheet("最终组队统计")
    # Each member: name | roster ID | student ID = 3 columns
    headers = ['队伍编号',
               '成员1（队长）', '成员1-ID', '成员1-学号',
               '成员2', '成员2-ID', '成员2-学号',
               '成员3', '成员3-ID', '成员3-学号',
               '人数', '选题方向', '数据备注']
    col_widths = [10,
                  14, 8, 16,
                  14, 8, 16,
                  14, 8, 16,
                  8, 42, 28]

    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'

    row = 2
    for i, team in enumerate(all_sorted_teams):
        fill = FILL_BY_SIZE.get(team['size'], FILL_WHITE)
        members = team['members']

        style_cell(ws, row, 1, i + 1, fill=fill)

        for mi in range(3):
            name_col = 2 + mi * 3       # 2, 5, 8
            roster_id_col = 3 + mi * 3  # 3, 6, 9
            stu_id_col = 4 + mi * 3     # 4, 7, 10
            if mi < len(members):
                name = members[mi]
                info = name_to_info.get(name, {})
                style_cell(ws, row, name_col, name, fill=fill)
                style_cell(ws, row, roster_id_col, info.get('roster_id', ''), FONT_SMALL, fill=fill)
                style_cell(ws, row, stu_id_col, info.get('student_id', ''), FONT_SMALL, fill=fill)
            else:
                style_cell(ws, row, name_col, '', fill=fill)
                style_cell(ws, row, roster_id_col, '', fill=fill)
                style_cell(ws, row, stu_id_col, '', fill=fill)

        style_cell(ws, row, 11, team['size'], FONT_BOLD, fill=fill)
        style_cell(ws, row, 12, team['topic'] if team['topic'] else '未确定', FONT_SMALL, fill=fill, align=LEFT)

        note_font = FONT_ORANGE if team['notes'] else FONT_SMALL
        style_cell(ws, row, 13, team['notes'], note_font, fill=fill, align=LEFT)

        row += 1

    # Summary row
    sizes = Counter(t['size'] for t in all_sorted_teams)
    total_people = sum(t['size'] for t in all_sorted_teams)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
    summary_text = (f"统计：三人组 {sizes.get(3, 0)} 队（{sizes.get(3, 0) * 3}人） | "
                    f"二人组 {sizes.get(2, 0)} 队（{sizes.get(2, 0) * 2}人） | "
                    f"一人组 {sizes.get(1, 0)} 队（{sizes.get(1, 0)}人） | "
                    f"共计 {len(all_sorted_teams)} 队（{total_people}人）")
    cell = ws.cell(row=row, column=1, value=summary_text)
    cell.font = FONT_BOLD
    cell.alignment = CENTER
    cell.fill = FILL_SUMMARY
    for c in range(1, 14):
        ws.cell(row=row, column=c).border = THIN_BORDER

    print(f"[Sheet1] 最终组队统计: {len(all_sorted_teams)} teams written")


# ============================================================
# Sheet 2: 未组队学生名单
# ============================================================
def generate_sheet2(wb, unpaired, name_to_info):
    """Sheet 2: Students not in any team."""
    ws = wb.create_sheet("未组队学生名单")
    headers = ['序号', '姓名', '学号', '当前年级', '当前院系', '行政班', '备注']
    col_widths = [8, 14, 16, 10, 22, 15, 16]

    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'

    # Sort by student ID
    sorted_unpaired = sorted(unpaired, key=lambda n: name_to_info.get(n, {}).get('student_id', ''))

    for i, name in enumerate(sorted_unpaired):
        row = i + 2
        info = name_to_info[name]
        style_cell(ws, row, 1, i + 1)
        style_cell(ws, row, 2, name)
        style_cell(ws, row, 3, info['student_id'], FONT_SMALL)
        grade_str = f"{int(info['grade'])}级" if info['grade'] else ''
        style_cell(ws, row, 4, grade_str, FONT_SMALL)
        style_cell(ws, row, 5, info['department'], FONT_SMALL)
        style_cell(ws, row, 6, info['class_name'], FONT_SMALL)
        note = '花名册有*号' if info['has_asterisk'] else ''
        style_cell(ws, row, 7, note, FONT_SMALL)

    # Summary
    row = len(sorted_unpaired) + 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=1, value=f"未组队人数：{len(sorted_unpaired)}人")
    cell.font = FONT_BOLD
    cell.alignment = CENTER
    cell.fill = FILL_SUMMARY
    for c in range(1, 8):
        ws.cell(row=row, column=c).border = THIN_BORDER

    print(f"[Sheet2] 未组队学生名单: {len(sorted_unpaired)} students")


# ============================================================
# Sheet 3: 数据分析
# ============================================================
def generate_sheet3(wb, all_sorted_teams, name_to_info, valid_names, unpaired):
    """Sheet 3: Data analysis with multiple sections."""
    ws = wb.create_sheet("数据分析")
    col_widths = [24, 14, 14, 20]
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = 'A2'

    row = 1
    total_students = len(valid_names)
    total_assigned = sum(t['size'] for t in all_sorted_teams)
    sizes = Counter(t['size'] for t in all_sorted_teams)

    # ── Section A: Team Size Distribution ──
    write_section_header(ws, row, 4, "【A】队伍规模分布")
    row += 1
    for c, h in enumerate(['规模', '队伍数', '占比', '学生数'], 1):
        style_cell(ws, row, c, h, FONT_HEADER, FILL_HEADER)
    row += 1

    size_labels = [(3, '三人组'), (2, '二人组'), (1, '一人组')]
    for sz, label in size_labels:
        count = sizes.get(sz, 0)
        pct = f"{count / len(all_sorted_teams) * 100:.1f}%" if all_sorted_teams else '0%'
        students = count * sz
        style_cell(ws, row, 1, label)
        style_cell(ws, row, 2, count)
        style_cell(ws, row, 3, pct)
        style_cell(ws, row, 4, students)
        row += 1

    # Total row
    style_cell(ws, row, 1, '合计', FONT_BOLD, FILL_SUMMARY)
    style_cell(ws, row, 2, len(all_sorted_teams), FONT_BOLD, FILL_SUMMARY)
    style_cell(ws, row, 3, '100%', FONT_BOLD, FILL_SUMMARY)
    style_cell(ws, row, 4, total_assigned, FONT_BOLD, FILL_SUMMARY)
    row += 2

    # ── Section B: Topic Direction Statistics ──
    write_section_header(ws, row, 4, "【B】选题方向统计")
    row += 1
    for c, h in enumerate(['方向类别', '队伍数', '占比', '人数'], 1):
        style_cell(ws, row, c, h, FONT_HEADER, FILL_HEADER)
    row += 1

    cat_counts = Counter()
    cat_people = Counter()
    for t in all_sorted_teams:
        cat = categorize_topic(t['topic'])
        cat_counts[cat] += 1
        cat_people[cat] += t['size']

    # Order: 未确定 last
    cat_order = [TOPIC_CAT_GAME, TOPIC_CAT_RENDER, TOPIC_CAT_AI, TOPIC_CAT_MOD, TOPIC_CAT_OTHER, TOPIC_CAT_UNKNOWN]
    for cat in cat_order:
        if cat not in cat_counts:
            continue
        count = cat_counts[cat]
        pct = f"{count / len(all_sorted_teams) * 100:.1f}%" if all_sorted_teams else '0%'
        style_cell(ws, row, 1, cat)
        style_cell(ws, row, 2, count)
        style_cell(ws, row, 3, pct)
        style_cell(ws, row, 4, cat_people[cat])
        row += 1

    style_cell(ws, row, 1, '合计', FONT_BOLD, FILL_SUMMARY)
    style_cell(ws, row, 2, len(all_sorted_teams), FONT_BOLD, FILL_SUMMARY)
    style_cell(ws, row, 3, '100%', FONT_BOLD, FILL_SUMMARY)
    style_cell(ws, row, 4, total_assigned, FONT_BOLD, FILL_SUMMARY)

    # Show individual topics
    row += 2
    write_section_header(ws, row, 4, "【B-2】具体选题方向明细")
    row += 1
    for c, h in enumerate(['队伍编号', '规模', '选题方向', '分类'], 1):
        style_cell(ws, row, c, h, FONT_HEADER, FILL_HEADER)
    row += 1

    for i, t in enumerate(all_sorted_teams):
        team_num = i + 1
        cat = categorize_topic(t['topic'])
        style_cell(ws, row, 1, team_num)
        style_cell(ws, row, 2, LABEL_BY_SIZE.get(t['size'], f"{t['size']}人"))
        style_cell(ws, row, 3, t['topic'] if t['topic'] else '未确定', FONT_SMALL, align=LEFT)
        style_cell(ws, row, 4, cat, FONT_SMALL)
        row += 1

    row += 1

    # ── Section C: Department Distribution ──
    write_section_header(ws, row, 4, "【C】院系分布（全体学生）")
    row += 1
    for c, h in enumerate(['院系', '人数', '占比', ''], 1):
        style_cell(ws, row, c, h, FONT_HEADER, FILL_HEADER)
    row += 1

    # Clean department names
    dept_counter = Counter()
    for name in valid_names:
        info = name_to_info[name]
        dept = info['department']
        # Normalize: "011计算机科学与技术系" → "计算机"
        if '计算机' in dept or '011' in dept:
            dept_counter['计算机'] += 1
        elif dept:
            dept_counter[dept] += 1
        else:
            dept_counter['未知'] += 1

    for dept, count in dept_counter.most_common():
        pct = f"{count / total_students * 100:.1f}%"
        style_cell(ws, row, 1, dept)
        style_cell(ws, row, 2, count)
        style_cell(ws, row, 3, pct)
        style_cell(ws, row, 4, '')
        row += 1

    style_cell(ws, row, 1, '合计', FONT_BOLD, FILL_SUMMARY)
    style_cell(ws, row, 2, total_students, FONT_BOLD, FILL_SUMMARY)
    style_cell(ws, row, 3, '100%', FONT_BOLD, FILL_SUMMARY)
    style_cell(ws, row, 4, '', FONT_BOLD, FILL_SUMMARY)
    row += 2

    # ── Section D: Grade Distribution ──
    write_section_header(ws, row, 4, "【D】年级分布（全体学生）")
    row += 1
    for c, h in enumerate(['年级', '人数', '占比', ''], 1):
        style_cell(ws, row, c, h, FONT_HEADER, FILL_HEADER)
    row += 1

    grade_counter = Counter()
    for name in valid_names:
        info = name_to_info[name]
        grade = info['grade']
        if grade:
            grade_counter[int(grade)] += 1
        else:
            grade_counter['未知'] += 1

    for grade in sorted(g for g in grade_counter if isinstance(g, int)):
        count = grade_counter[grade]
        pct = f"{count / total_students * 100:.1f}%"
        style_cell(ws, row, 1, f"{grade}级")
        style_cell(ws, row, 2, count)
        style_cell(ws, row, 3, pct)
        style_cell(ws, row, 4, '')
        row += 1
    if '未知' in grade_counter:
        style_cell(ws, row, 1, '未知', FONT_SMALL)
        style_cell(ws, row, 2, grade_counter['未知'])
        style_cell(ws, row, 3, f"{grade_counter['未知'] / total_students * 100:.1f}%")
        style_cell(ws, row, 4, '')
        row += 1

    style_cell(ws, row, 1, '合计', FONT_BOLD, FILL_SUMMARY)
    style_cell(ws, row, 2, total_students, FONT_BOLD, FILL_SUMMARY)
    style_cell(ws, row, 3, '100%', FONT_BOLD, FILL_SUMMARY)
    style_cell(ws, row, 4, '', FONT_BOLD, FILL_SUMMARY)
    row += 2

    # ── Section E: Overall Summary ──
    write_section_header(ws, row, 4, "【E】总体摘要")
    row += 1

    summaries = [
        ('总学生数', total_students),
        ('已组队学生数', f"{total_assigned} ({total_assigned / total_students * 100:.1f}%)"),
        ('未组队学生数', f"{len(unpaired)} ({len(unpaired) / total_students * 100:.1f}%)"),
        ('总队伍数', len(all_sorted_teams)),
        ('平均每队人数', f"{total_assigned / len(all_sorted_teams):.1f}" if all_sorted_teams else 'N/A'),
        ('三人组队伍数', sizes.get(3, 0)),
        ('二人组队伍数', sizes.get(2, 0)),
        ('一人组队伍数', sizes.get(1, 0)),
    ]
    for label, value in summaries:
        style_cell(ws, row, 1, label, FONT_BOLD, align=LEFT)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        style_cell(ws, row, 2, str(value), align=LEFT)
        row += 1

    print(f"[Sheet3] 数据分析: 5 sections written")


# ============================================================
# Sheet 4: 全部学生分配情况
# ============================================================
def generate_sheet4(wb, all_sorted_teams, name_to_info, valid_names, unpaired):
    """Sheet 4: All students with team assignment status."""
    ws = wb.create_sheet("全部学生分配情况")
    headers = ['序号', '姓名', '学号', '所在队伍编号', '队伍规模', '状态', '当前年级', '当前院系']
    col_widths = [8, 14, 16, 14, 10, 10, 10, 22]

    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'

    # Build team mapping
    team_map = {}
    for i, team in enumerate(all_sorted_teams):
        tn = i + 1
        for m in team['members']:
            team_map[m] = (tn, team['size'])

    # Sort by roster ID
    sorted_names = sorted(valid_names, key=lambda n: name_to_info.get(n, {}).get('roster_id', 0))

    for i, name in enumerate(sorted_names):
        row = i + 2
        info = name_to_info[name]

        style_cell(ws, row, 1, info['roster_id'])
        style_cell(ws, row, 2, name)
        style_cell(ws, row, 3, info['student_id'], FONT_SMALL)
        grade_str = f"{int(info['grade'])}级" if info['grade'] else ''
        style_cell(ws, row, 7, grade_str, FONT_SMALL)
        style_cell(ws, row, 8, info['department'], FONT_SMALL, align=LEFT)

        if name in team_map:
            tn, ts = team_map[name]
            fill = FILL_BY_SIZE.get(ts, FILL_WHITE)
            style_cell(ws, row, 4, tn, fill=fill)
            style_cell(ws, row, 5, LABEL_BY_SIZE.get(ts, f'{ts}人'), fill=fill)
            style_cell(ws, row, 6, '已组队', FONT_GREEN, fill=fill)
        else:
            style_cell(ws, row, 4, '-')
            style_cell(ws, row, 5, '-')
            style_cell(ws, row, 6, '未组队', FONT_RED, FILL_RED_LIGHT)

    # Summary
    row = len(sorted_names) + 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    assigned_count = len(valid_names) - len(unpaired)
    cell = ws.cell(row=row, column=1,
                   value=f"总学生 {len(valid_names)}人 | 已组队 {assigned_count}人 | 未组队 {len(unpaired)}人")
    cell.font = FONT_BOLD
    cell.alignment = CENTER
    cell.fill = FILL_SUMMARY
    for c in range(1, 9):
        ws.cell(row=row, column=c).border = THIN_BORDER

    print(f"[Sheet4] 全部学生分配情况: {len(sorted_names)} students")


# ============================================================
# STEP 5: Main
# ============================================================
def main():
    print("=" * 60)
    print("CG26 Final Team Formation Statistics Generator")
    print("=" * 60)

    # Load data
    name_to_info, valid_names = load_roster(ROSTER_PATH)
    raw_teams = load_teams(TEAMS_PATH)

    # Validate & cross-reference
    validated_teams, invalid_names, student_assigned = validate_teams(raw_teams, valid_names)

    # Find unpaired
    unpaired = compute_unpaired(valid_names, student_assigned)

    # Sort: 3-person → 2-person → 1-person
    teams_3 = [t for t in validated_teams if t['size'] == 3]
    teams_2 = [t for t in validated_teams if t['size'] == 2]
    teams_1 = [t for t in validated_teams if t['size'] == 1]
    all_sorted_teams = teams_3 + teams_2 + teams_1

    print(f"\n{'=' * 40}")
    print(f"SUMMARY")
    print(f"{'=' * 40}")
    print(f"  Total students in roster: {len(valid_names)}")
    print(f"  Total students in teams:  {len(student_assigned)}")
    print(f"  Unpaired students:        {len(unpaired)}")
    print(f"  3-person teams:           {len(teams_3)}")
    print(f"  2-person teams:           {len(teams_2)}")
    print(f"  1-person teams:           {len(teams_1)}")
    print(f"  Total teams:              {len(all_sorted_teams)}")

    # Generate Excel
    print(f"\n{'=' * 40}")
    print(f"GENERATING EXCEL OUTPUT")
    print(f"{'=' * 40}")

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    generate_sheet1(wb, all_sorted_teams, name_to_info)
    generate_sheet2(wb, unpaired, name_to_info)
    generate_sheet3(wb, all_sorted_teams, name_to_info, valid_names, unpaired)
    generate_sheet4(wb, all_sorted_teams, name_to_info, valid_names, unpaired)

    # Save
    wb.save(OUTPUT_XLSX)
    print(f"\n[OUTPUT] Excel saved to: {OUTPUT_XLSX}")

    # Also save JSON summary
    summary = {
        "total_students_in_roster": len(valid_names),
        "total_assigned": len(student_assigned),
        "unpaired_count": len(unpaired),
        "teams_3_person": len(teams_3),
        "teams_2_person": len(teams_2),
        "teams_1_person": len(teams_1),
        "total_teams": len(all_sorted_teams),
        "invalid_names_in_teams": sorted(invalid_names),
        "unpaired_names": sorted(unpaired),
        "teams": [
            {
                "team_number": i + 1,
                "members": t['members'],
                "size": t['size'],
                "topic": t['topic'],
                "notes": t['notes'],
            }
            for i, t in enumerate(all_sorted_teams)
        ],
        "topic_distribution": {
            cat: count for cat, count in
            Counter(categorize_topic(t['topic']) for t in all_sorted_teams).most_common()
        },
    }
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    print(f"[OUTPUT] JSON summary saved to: {OUTPUT_JSON}")

    print(f"\n{'=' * 60}")
    print(f"DONE. Sheets generated:")
    print(f"  1. 最终组队统计 - {len(all_sorted_teams)} teams")
    print(f"  2. 未组队学生名单 - {len(unpaired)} students")
    print(f"  3. 数据分析 - 5 sections")
    print(f"  4. 全部学生分配情况 - {len(valid_names)} students")
    print(f"{'=' * 60}")


if __name__ == '__main__':
    main()
