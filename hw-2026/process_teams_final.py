# -*- coding: utf-8 -*-
"""
Final team formation processing script.
1. Read student roster
2. Read team formation stats
3. Filter invalid names
4. Create final Excel table sorted 3-person → 2-person → 1-person
5. List unmatched students
"""
import pandas as pd
import json
from collections import defaultdict

# ============================================================
# STEP 1: Read student roster
# ============================================================
df_students = pd.read_excel(r'D:\Projects\USTC-CG-Score\hw-2026\StudentNames-new.xlsx')

# Columns: ID号, 学号, 姓名, 当前年级, 当前院系, 行政班
valid_students = df_students[df_students['姓名'].notna()].copy()
valid_students = valid_students.dropna(subset=['学号'])
valid_students['ID号'] = valid_students['ID号'].astype(int)

# Build name → info mapping (strip trailing * from roster names)
name_to_info = {}
roster_name_to_clean = {}  # roster name → cleaned name
for _, row in valid_students.iterrows():
    roster_name = str(row['姓名']).strip()
    # Some names in roster have trailing "*" - strip it for matching
    clean_name = roster_name.rstrip('*')
    name_to_info[clean_name] = {
        "roster_id": int(row['ID号']),
        "student_id": str(row['学号']),
        "roster_name": roster_name,
        "grade": row['当前年级'] if pd.notna(row['当前年级']) else '',
        "department": str(row['当前院系']).strip() if pd.notna(row['当前院系']) else '',
        "class": str(row['行政班']).strip() if pd.notna(row['行政班']) else ''
    }
    roster_name_to_clean[roster_name] = clean_name

valid_names_set = set(name_to_info.keys())
print(f"Total valid students in roster: {len(valid_names_set)}")

# ============================================================
# STEP 2: Read team formation stats
# ============================================================
df_teams = pd.read_excel(r'D:\Projects\USTC-CG-Score\hw-2026\CG2026FinalProject组队统计表（收集结果）.xlsx')

# Columns: 提交者（自动）, 提交时间（自动）, 队长姓名（必填）, 队员姓名 1, 队员姓名 2, 选题
raw_teams = []
all_submitted_names = set()

for idx, row in df_teams.iterrows():
    captain = str(row['队长姓名（必填）']).strip() if pd.notna(row['队长姓名（必填）']) else ''
    m1_raw = str(row['队员姓名 1']).strip() if pd.notna(row['队员姓名 1']) else ''
    m2_raw = str(row['队员姓名 2']).strip() if pd.notna(row['队员姓名 2']) else ''
    topic = str(row['选题']).strip() if pd.notna(row['选题']) else ''
    
    # Filter out placeholder values
    def is_placeholder(s):
        return s in ('', 'nan', '无', '空', 'None', 'NaN', '已删除')
    
    members_raw = [captain]
    if not is_placeholder(m1_raw):
        members_raw.append(m1_raw)
    if not is_placeholder(m2_raw):
        members_raw.append(m2_raw)
    
    all_submitted_names.update(members_raw)
    
    raw_teams.append({
        'row': idx,
        'captain': captain,
        'members_raw': members_raw,
        'topic': topic if topic and topic != 'nan' else ''
    })

print(f"Total submissions: {len(raw_teams)}")
print(f"Total unique names submitted: {len(all_submitted_names)}")

# ============================================================
# STEP 3: Validate names against roster
# ============================================================
invalid_names = all_submitted_names - valid_names_set
valid_submitted_names = all_submitted_names & valid_names_set

print(f"\n=== INVALID NAMES (not in roster) ===")
for n in sorted(invalid_names):
    print(f"  [INVALID] {n}")

print(f"\n=== VALID NAMES (in roster) ===")
for n in sorted(valid_submitted_names):
    print(f"  [VALID] {n}")

# ============================================================
# STEP 4: Build validated teams
# ============================================================
# Track which students are assigned to which team
student_team_map = {}  # name → team_index
validated_teams = []

for rt in raw_teams:
    valid_members = [m for m in rt['members_raw'] if m in valid_names_set]
    
    # Deduplicate within the team (some names might appear twice)
    seen = set()
    deduped = []
    for m in valid_members:
        if m not in seen:
            seen.add(m)
            deduped.append(m)
    valid_members = deduped
    
    if len(valid_members) == 0:
        # Team has no valid members at all
        print(f"\n  Row {rt['row']}: All members invalid! Raw: {rt['members_raw']}")
        continue
    
    validated_teams.append({
        'row': rt['row'],
        'members': valid_members,
        'topic': rt['topic'],
        'raw_captain': rt['captain']
    })

# ============================================================
# STEP 5: Handle duplicate students across teams
# ============================================================
# A student might appear in multiple team submissions.
# We need to handle this. Strategy: keep the first occurrence, remove from later teams.

print("\n=== DUPLICATE CHECK ===")
dedup_teams = []
student_assigned = set()

for vt in validated_teams:
    cleaned_members = []
    for m in vt['members']:
        if m in student_assigned:
            print(f"  [WARN] {m} already in another team, removing from row {vt['row']}")
        else:
            cleaned_members.append(m)
            student_assigned.add(m)
    
    if len(cleaned_members) > 0:
        dedup_teams.append({
            'members': cleaned_members,
            'topic': vt['topic'],
            'raw_captain': vt['raw_captain']
        })

# ============================================================
# STEP 6: Sort: 3-person → 2-person → 1-person
# ============================================================
teams_3 = [t for t in dedup_teams if len(t['members']) == 3]
teams_2 = [t for t in dedup_teams if len(t['members']) == 2]
teams_1 = [t for t in dedup_teams if len(t['members']) == 1]

print(f"\n=== TEAM SUMMARY ===")
print(f"3-person teams: {len(teams_3)}")
print(f"2-person teams: {len(teams_2)}")
print(f"1-person teams: {len(teams_1)}")
print(f"Total teams: {len(teams_3) + len(teams_2) + len(teams_1)}")

# ============================================================
# STEP 7: Find unmatched students (not in any team)
# ============================================================
all_assigned_names = set()
for t in dedup_teams:
    all_assigned_names.update(t['members'])

unmatched_names = valid_names_set - all_assigned_names
print(f"\n=== UNMATCHED STUDENTS ({len(unmatched_names)}) ===")
for n in sorted(unmatched_names):
    info = name_to_info[n]
    print(f"  {info['roster_id']:3d}: {info['student_id']} - {n}")

# ============================================================
# STEP 8: Create output Excel
# ============================================================
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# --- Sheet 1: 最终组队统计 ---
ws1 = wb.active
ws1.title = "最终组队统计"

# Styles
header_font = Font(name='微软雅黑', bold=True, size=11)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font_white = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
cell_font = Font(name='微软雅黑', size=11)
cell_font_small = Font(name='微软雅黑', size=10)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Team count colors
fill_3 = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Green for 3-person
fill_2 = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')  # Blue for 2-person
fill_1 = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')  # Orange for 1-person

# Headers
headers = ['队伍编号', '成员1（队长）', '成员1-学号', '成员2', '成员2-学号', '成员3', '成员3-学号', '人数', '选题方向']
for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# Column widths
col_widths = [10, 14, 16, 14, 16, 14, 16, 8, 40]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Sort all teams: 3-person → 2-person → 1-person
all_sorted_teams = teams_3 + teams_2 + teams_1

# Assign team numbers
team_number = 1
row_idx = 2

for team in all_sorted_teams:
    members = team['members']
    
    # Determine fill color based on size
    if len(members) == 3:
        fill = fill_3
    elif len(members) == 2:
        fill = fill_2
    else:
        fill = fill_1
    
    # Team number
    cell = ws1.cell(row=row_idx, column=1, value=team_number)
    cell.font = cell_font
    cell.alignment = center_align
    cell.border = thin_border
    
    # Member 1 (captain position or first member)
    m1_info = name_to_info.get(members[0], {})
    cell = ws1.cell(row=row_idx, column=2, value=members[0])
    cell.font = cell_font
    cell.alignment = center_align
    cell.border = thin_border
    cell.fill = fill
    
    cell = ws1.cell(row=row_idx, column=3, value=m1_info.get('student_id', ''))
    cell.font = cell_font_small
    cell.alignment = center_align
    cell.border = thin_border
    cell.fill = fill
    
    # Member 2
    if len(members) >= 2:
        m2_info = name_to_info.get(members[1], {})
        cell = ws1.cell(row=row_idx, column=4, value=members[1])
        cell.font = cell_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = fill
        cell = ws1.cell(row=row_idx, column=5, value=m2_info.get('student_id', ''))
        cell.font = cell_font_small
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = fill
    else:
        cell = ws1.cell(row=row_idx, column=4, value='')
        cell.border = thin_border
        cell.fill = fill
        cell = ws1.cell(row=row_idx, column=5, value='')
        cell.border = thin_border
        cell.fill = fill
    
    # Member 3
    if len(members) >= 3:
        m3_info = name_to_info.get(members[2], {})
        cell = ws1.cell(row=row_idx, column=6, value=members[2])
        cell.font = cell_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = fill
        cell = ws1.cell(row=row_idx, column=7, value=m3_info.get('student_id', ''))
        cell.font = cell_font_small
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = fill
    else:
        cell = ws1.cell(row=row_idx, column=6, value='')
        cell.border = thin_border
        cell.fill = fill
        cell = ws1.cell(row=row_idx, column=7, value='')
        cell.border = thin_border
        cell.fill = fill
    
    # Team size
    cell = ws1.cell(row=row_idx, column=8, value=len(members))
    cell.font = cell_font
    cell.alignment = center_align
    cell.border = thin_border
    cell.fill = fill
    
    # Topic
    cell = ws1.cell(row=row_idx, column=9, value=team['topic'] if team['topic'] else '未确定')
    cell.font = cell_font_small
    cell.alignment = left_align
    cell.border = thin_border
    cell.fill = fill
    
    team_number += 1
    row_idx += 1

# Add summary rows
row_idx += 1
summary_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
for col in range(1, 10):
    cell = ws1.cell(row=row_idx, column=col)
    cell.fill = summary_fill
    cell.border = thin_border
ws1.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)
cell = ws1.cell(row=row_idx, column=1, 
    value=f"统计：三人组 {len(teams_3)} 队（{len(teams_3)*3}人）| 二人组 {len(teams_2)} 队（{len(teams_2)*2}人）| 一人组 {len(teams_1)} 队（{len(teams_1)*1}人）| 共计 {len(all_sorted_teams)} 队（{len(teams_3)*3 + len(teams_2)*2 + len(teams_1)}人）")
cell.font = Font(name='微软雅黑', bold=True, size=11)
cell.alignment = center_align

# --- Sheet 2: 未完成组队 ---
ws2 = wb.create_sheet("未完成组队名单")

headers2 = ['序号', '姓名', '学号', '当前年级', '当前院系', '行政班']
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

col_widths2 = [8, 14, 16, 10, 20, 15]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

unmatched_sorted = sorted(unmatched_names)  # Sort by name for now
for i, name in enumerate(unmatched_sorted, 1):
    info = name_to_info[name]
    row_num = i + 1
    
    cell = ws2.cell(row=row_num, column=1, value=info['roster_id'])
    cell.font = cell_font
    cell.alignment = center_align
    cell.border = thin_border
    
    cell = ws2.cell(row=row_num, column=2, value=name)
    cell.font = cell_font
    cell.alignment = center_align
    cell.border = thin_border
    
    cell = ws2.cell(row=row_num, column=3, value=info['student_id'])
    cell.font = cell_font_small
    cell.alignment = center_align
    cell.border = thin_border
    
    cell = ws2.cell(row=row_num, column=4, value=f"{int(info['grade'])}级" if info['grade'] else '')
    cell.font = cell_font_small
    cell.alignment = center_align
    cell.border = thin_border
    
    cell = ws2.cell(row=row_num, column=5, value=info['department'])
    cell.font = cell_font_small
    cell.alignment = center_align
    cell.border = thin_border
    
    cell = ws2.cell(row=row_num, column=6, value=info['class'])
    cell.font = cell_font_small
    cell.alignment = center_align
    cell.border = thin_border

# Unmatched summary
summary_row = len(unmatched_sorted) + 3
for col in range(1, 7):
    cell = ws2.cell(row=summary_row, column=col)
    cell.fill = summary_fill
    cell.border = thin_border
ws2.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=6)
cell = ws2.cell(row=summary_row, column=1, value=f"未组队人数：{len(unmatched_sorted)}人")
cell.font = Font(name='微软雅黑', bold=True, size=11)
cell.alignment = center_align

# --- Sheet 3: 无效名字记录 ---
ws3 = wb.create_sheet("非法名字记录")

headers3 = ['序号', '原始提交中的名字', '所在行', '提交者', '备注']
for col, h in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=col, value=h)
    cell.font = header_font_white
    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    cell.alignment = center_align
    cell.border = thin_border

col_widths3 = [8, 20, 10, 20, 30]
for i, w in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

invalid_count = 0
for rt in raw_teams:
    for m in rt['members_raw']:
        if m in invalid_names:
            invalid_count += 1
            row_num = invalid_count + 1
            cell = ws3.cell(row=row_num, column=1, value=invalid_count)
            cell.font = cell_font
            cell.alignment = center_align
            cell.border = thin_border
            
            cell = ws3.cell(row=row_num, column=2, value=m)
            cell.font = cell_font
            cell.alignment = center_align
            cell.border = thin_border
            
            cell = ws3.cell(row=row_num, column=3, value=rt['row'] + 1)
            cell.font = cell_font
            cell.alignment = center_align
            cell.border = thin_border
            
            cell = ws3.cell(row=row_num, column=4, value=rt['captain'])
            cell.font = cell_font
            cell.alignment = center_align
            cell.border = thin_border
            
            cell = ws3.cell(row=row_num, column=5, value='不在学生名单中')
            cell.font = cell_font
            cell.alignment = center_align
            cell.border = thin_border

# --- Sheet 4: 全部学生分配情况 ---
ws4 = wb.create_sheet("全部学生分配情况")

headers4 = ['序号', '姓名', '学号', '所在队伍编号', '队伍人数', '状态']
for col, h in enumerate(headers4, 1):
    cell = ws4.cell(row=1, column=col, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

col_widths4 = [8, 14, 16, 14, 10, 12]
for i, w in enumerate(col_widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

team_num_map = {}
for i, team in enumerate(all_sorted_teams):
    tn = i + 1
    for m in team['members']:
        team_num_map[m] = (tn, len(team['members']))

all_names_sorted = sorted(valid_names_set)
for i, name in enumerate(all_names_sorted):
    row_num = i + 2
    info = name_to_info[name]
    
    cell = ws4.cell(row=row_num, column=1, value=info['roster_id'])
    cell.font = cell_font
    cell.alignment = center_align
    cell.border = thin_border
    
    cell = ws4.cell(row=row_num, column=2, value=name)
    cell.font = cell_font
    cell.alignment = center_align
    cell.border = thin_border
    
    cell = ws4.cell(row=row_num, column=3, value=info['student_id'])
    cell.font = cell_font_small
    cell.alignment = center_align
    cell.border = thin_border
    
    if name in team_num_map:
        tn, ts = team_num_map[name]
        cell = ws4.cell(row=row_num, column=4, value=tn)
        cell.font = cell_font
        cell.alignment = center_align
        cell.border = thin_border
        
        cell = ws4.cell(row=row_num, column=5, value=f"{ts}人组")
        cell.font = cell_font
        cell.alignment = center_align
        cell.border = thin_border
        
        cell = ws4.cell(row=row_num, column=6, value='已组队')
        cell.font = Font(name='微软雅黑', size=11, color='006100')
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = fill_3 if ts == 3 else (fill_2 if ts == 2 else fill_1)
    else:
        cell = ws4.cell(row=row_num, column=4, value='-')
        cell.font = cell_font
        cell.alignment = center_align
        cell.border = thin_border
        
        cell = ws4.cell(row=row_num, column=5, value='-')
        cell.font = cell_font
        cell.alignment = center_align
        cell.border = thin_border
        
        cell = ws4.cell(row=row_num, column=6, value='未组队')
        cell.font = Font(name='微软雅黑', size=11, color='9C0006')
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

# Freeze header rows
ws1.freeze_panes = 'A2'
ws2.freeze_panes = 'A2'
ws3.freeze_panes = 'A2'
ws4.freeze_panes = 'A2'

# Auto-fit by setting row height
ws1.row_dimensions[1].height = 30
ws2.row_dimensions[1].height = 30
ws3.row_dimensions[1].height = 30
ws4.row_dimensions[1].height = 30

# Save
output_path = r'D:\Projects\USTC-CG-Score\hw-2026\CG2026FinalProject组队统计（最终版）.xlsx'
wb.save(output_path)

print(f"\n{'='*60}")
print(f"OUTPUT SAVED TO: {output_path}")
print(f"{'='*60}")
print(f"Sheet 1 - 最终组队统计: {len(all_sorted_teams)} teams")
print(f"Sheet 2 - 未完成组队名单: {len(unmatched_names)} students")
print(f"Sheet 3 - 非法名字记录: {invalid_count} invalid entries")
print(f"Sheet 4 - 全部学生分配情况: {len(all_names_sorted)} students")

# Also save a summary JSON for reference
summary = {
    "total_students": len(valid_names_set),
    "teams_3_person": len(teams_3),
    "teams_2_person": len(teams_2),
    "teams_1_person": len(teams_1),
    "total_teams": len(all_sorted_teams),
    "total_assigned": len(all_assigned_names),
    "unmatched_count": len(unmatched_names),
    "unmatched_names": sorted(unmatched_names),
    "invalid_names": sorted(invalid_names),
    "teams": [
        {
            "team_number": i + 1,
            "members": t['members'],
            "size": len(t['members']),
            "topic": t['topic']
        }
        for i, t in enumerate(all_sorted_teams)
    ]
}
with open(r'D:\Projects\USTC-CG-Score\hw-2026\team_summary.json', 'w', encoding='utf-8') as f:
    json.dump(summary, f, ensure_ascii=False, indent=2)
print(f"Summary JSON saved to: team_summary.json")
